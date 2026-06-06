"""Extractor de texto para documentos Excel (.xlsx).

Devuelve una representación Markdown-like de todas las hojas: cada hoja se
introduce con un encabezado y sus filas se formatean como tabla con pipes.

Este formato facilita que el modelo textual identifique columnas y filas de una
rúbrica sin necesidad de mantener su estructura tabular original.
"""

from pathlib import Path

from openpyxl import load_workbook


def extract_xlsx(path: str | Path) -> str:
    """Extrae el contenido de un .xlsx como texto Markdown-like.

    Recorre todas las hojas del libro. Para cada hoja escribe un encabezado
    ## <nombre de la hoja> y a continuación las filas no vacías formateadas
    como | celda1 | celda2 | ... |. Las celdas vacías se rellenan con cadena
    vacía para preservar la alineación de columnas.
    """
    workbook = load_workbook(filename=str(path), data_only=True, read_only=True)

    sections: list[str] = []
    for sheet in workbook.worksheets:
        rows = _format_rows(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        sections.append(f"## {sheet.title}\n\n" + "\n".join(rows))

    workbook.close()
    return "\n\n".join(sections)


def _format_rows(rows_iter) -> list[str]:
    formatted: list[str] = []
    for row in rows_iter:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        cells = [_format_cell(cell) for cell in row]
        formatted.append("| " + " | ".join(cells) + " |")
    return formatted


def _format_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()
